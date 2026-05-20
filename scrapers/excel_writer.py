"""Gera o Excel com múltiplas abas (uma por cenário) + abas de apoio."""
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
TITULO_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
SUBTITULO_FONT = Font(name="Arial", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _formatar_cabecalho(ws, linha, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=linha, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _ajustar_larguras(ws, larguras):
    for col_letra, largura in larguras.items():
        ws.column_dimensions[col_letra].width = largura


def gerar_aba_cenario(wb, nome_cenario, agregacao, pesquisas_brutas, pesos_institutos, config):
    nome_aba = nome_cenario.replace("/", "-").replace(":", "-")[:31]
    ws = wb.create_sheet(nome_aba)

    ws["A1"] = f"Cenário: {nome_cenario}"
    ws["A1"].font = TITULO_FONT
    ws.merge_cells("A1:F1")

    ws["A3"] = "Data de referência:"
    ws["B3"] = agregacao["data_referencia"].isoformat()
    ws["A4"] = "Pesquisas consideradas:"
    ws["B4"] = agregacao["n_pesquisas"]
    for cell in ["A3", "A4"]:
        ws[cell].font = SUBTITULO_FONT

    candidatos = agregacao["candidatos"]

    ws["A6"] = "MÉDIA PONDERADA"
    ws["A6"].font = SUBTITULO_FONT
    ws["A7"] = "Candidato"
    ws["B7"] = "Média (%)"
    _formatar_cabecalho(ws, 7, 2)

    for i, candidato in enumerate(candidatos, start=8):
        ws.cell(row=i, column=1, value=candidato).font = Font(name="Arial")
        c = ws.cell(row=i, column=2, value=round(agregacao["medias"].get(candidato, 0), 2))
        c.font = Font(name="Arial")
        c.number_format = "0.00"
        c.alignment = Alignment(horizontal="right")

    linha_inicio_lista = 9 + len(candidatos) + 2
    ws.cell(row=linha_inicio_lista - 1, column=1, value="PESQUISAS UTILIZADAS").font = SUBTITULO_FONT

    df = agregacao["detalhamento"]
    if not df.empty:
        cols = ["instituto", "data_fim_campo", "amostra", "peso_final"] + candidatos
        df = df[cols].copy().sort_values("data_fim_campo", ascending=False)

        for j, col in enumerate(cols, start=1):
            ws.cell(row=linha_inicio_lista, column=j, value=col)
        _formatar_cabecalho(ws, linha_inicio_lista, len(cols))

        for i, row in enumerate(df.itertuples(index=False), start=linha_inicio_lista + 1):
            for j, valor in enumerate(row, start=1):
                c = ws.cell(row=i, column=j, value=valor)
                c.font = Font(name="Arial")
                c.border = THIN_BORDER
                if isinstance(valor, float):
                    c.number_format = "0.00"

    larguras = {"A": 22, "B": 16}
    for j in range(3, len(candidatos) + 5):
        larguras[get_column_letter(j)] = 14
    _ajustar_larguras(ws, larguras)


def gerar_aba_evolucao(wb, nome_cenario, serie, top_n=6):
    """
    Cria uma aba com a série temporal (média móvel ponderada) e um gráfico
    de linha mostrando a evolução dos principais candidatos ao longo do tempo.

    `serie` é o dict retornado por ponderacao.agregar_serie_temporal().
    """
    nome_aba = f"Evolução {nome_cenario}".replace("/", "-").replace(":", "-")[:31]
    ws = wb.create_sheet(nome_aba)

    ws["A1"] = f"Evolução temporal — {nome_cenario}"
    ws["A1"].font = TITULO_FONT
    ws.merge_cells("A1:H1")

    pontos = serie.get("pontos", [])
    if not pontos:
        ws["A3"] = "(sem dados para gerar a série temporal)"
        ws["A3"].font = Font(name="Arial", italic=True, color="888888")
        return

    ws["A3"] = f"Janela: {serie['janela_dias']} dias | Passo: {serie['passo_dias']} dias"
    ws["A3"].font = Font(name="Arial", italic=True, size=10, color="666666")

    # Escolher os top_n candidatos pela média mais recente com dados
    medias_recentes = {}
    for pt in reversed(pontos):
        if pt["n_pesquisas"] > 0:
            medias_recentes = pt["medias"]
            break
    especiais = {"Outros", "Branco/Nulo", "Não sabe"}
    candidatos_ordenados = sorted(
        [c for c in serie["candidatos"] if c not in especiais],
        key=lambda c: medias_recentes.get(c, 0),
        reverse=True,
    )
    candidatos_grafico = candidatos_ordenados[:top_n]

    # Tabela: linha de cabeçalho (Data + candidatos)
    linha_cab = 5
    ws.cell(row=linha_cab, column=1, value="Data")
    ws.cell(row=linha_cab, column=2, value="Nº pesq.")
    for j, cand in enumerate(candidatos_grafico, start=3):
        ws.cell(row=linha_cab, column=j, value=cand)
    _formatar_cabecalho(ws, linha_cab, 2 + len(candidatos_grafico))

    # Linhas de dados
    primeira_dado = linha_cab + 1
    linha = primeira_dado
    for pt in pontos:
        ws.cell(row=linha, column=1, value=pt["data"].isoformat()).font = Font(name="Arial")
        ws.cell(row=linha, column=2, value=pt["n_pesquisas"]).font = Font(name="Arial")
        for j, cand in enumerate(candidatos_grafico, start=3):
            val = pt["medias"].get(cand)
            c = ws.cell(row=linha, column=j, value=round(val, 2) if val else None)
            c.font = Font(name="Arial")
            if val:
                c.number_format = "0.00"
        linha += 1
    ultima_dado = linha - 1

    # Gráfico de linha
    chart = LineChart()
    chart.title = f"Intenção de voto — {nome_cenario}"
    chart.style = 12
    chart.y_axis.title = "% intenção de voto"
    chart.x_axis.title = "Data"
    chart.height = 10
    chart.width = 22

    dados = Reference(ws, min_col=3, max_col=2 + len(candidatos_grafico),
                      min_row=linha_cab, max_row=ultima_dado)
    categorias = Reference(ws, min_col=1, min_row=primeira_dado, max_row=ultima_dado)
    chart.add_data(dados, titles_from_data=True)
    chart.set_categories(categorias)

    # Linhas mais grossas e suaves
    for serie_graf in chart.series:
        serie_graf.smooth = True
        serie_graf.graphicalProperties.line.width = 22000  # ~1.7pt

    col_grafico = get_column_letter(2 + len(candidatos_grafico) + 2)
    ws.add_chart(chart, f"{col_grafico}5")

    # Larguras
    larguras = {"A": 12, "B": 10}
    for j in range(3, 3 + len(candidatos_grafico)):
        larguras[get_column_letter(j)] = 12
    _ajustar_larguras(ws, larguras)


def gerar_aba_todas_pesquisas(wb, todas_pesquisas):
    ws = wb.create_sheet("Todas as Pesquisas")

    if not todas_pesquisas:
        ws["A1"] = "Nenhuma pesquisa no banco."
        return

    candidatos_set = set()
    for p in todas_pesquisas:
        candidatos_set.update(p.get("resultados", {}).keys())
    especiais = ["Outros", "Branco/Nulo", "Não sabe"]
    candidatos = sorted([c for c in candidatos_set if c not in especiais])
    candidatos += [c for c in especiais if c in candidatos_set]

    cols_meta = [
        "cenario", "instituto", "contratante", "data_inicio_campo",
        "data_fim_campo", "amostra", "margem_erro", "turno",
        "metodologia", "votos_validos", "registro_tse", "url_fonte",
    ]
    cols = cols_meta + candidatos

    for j, col in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=col)
    _formatar_cabecalho(ws, 1, len(cols))

    for i, p in enumerate(todas_pesquisas, start=2):
        for j, col in enumerate(cols_meta, start=1):
            c = ws.cell(row=i, column=j, value=p.get(col))
            c.font = Font(name="Arial")
            c.border = THIN_BORDER
        for j, cand in enumerate(candidatos, start=len(cols_meta) + 1):
            val = p.get("resultados", {}).get(cand)
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial")
            c.border = THIN_BORDER
            if val is not None:
                c.number_format = "0.00"

    larguras = {get_column_letter(j): 14 for j in range(1, len(cols) + 1)}
    larguras["A"] = 22
    larguras["B"] = 18
    larguras["C"] = 18
    larguras["L"] = 16
    _ajustar_larguras(ws, larguras)


def gerar_aba_config(wb, config, pesos_institutos):
    ws = wb.create_sheet("Configuração")
    ws["A1"] = "Pesos por instituto"
    ws["A1"].font = TITULO_FONT

    ws["A3"] = "Instituto"
    ws["B3"] = "Peso"
    _formatar_cabecalho(ws, 3, 2)

    linha = 4
    for inst, peso in sorted(pesos_institutos.items()):
        ws.cell(row=linha, column=1, value=inst).font = Font(name="Arial")
        c = ws.cell(row=linha, column=2, value=peso)
        c.font = Font(name="Arial")
        c.number_format = "0.00"
        linha += 1

    linha += 2
    ws.cell(row=linha, column=1, value="Parâmetros da ponderação").font = TITULO_FONT
    linha += 2
    ws.cell(row=linha, column=1, value="Half-life (dias)")
    ws.cell(row=linha, column=2, value=config["ponderacao"]["recencia"]["half_life_dias"])
    linha += 1
    ws.cell(row=linha, column=1, value="Amostra de referência")
    ws.cell(row=linha, column=2, value=config["ponderacao"]["amostra"]["amostra_referencia"])
    linha += 1
    ws.cell(row=linha, column=1, value="Janela de pesquisas (dias)")
    ws.cell(row=linha, column=2, value=config["ponderacao"]["janela_dias"])

    linha += 2
    ws.cell(row=linha, column=1, value="Cenários ativos").font = TITULO_FONT
    linha += 2
    ws.cell(row=linha, column=1, value="Cenário")
    ws.cell(row=linha, column=2, value="Turno")
    ws.cell(row=linha, column=3, value="Candidatos exigidos")
    _formatar_cabecalho(ws, linha, 3)
    linha += 1
    for cen in config["cenarios"]:
        ws.cell(row=linha, column=1, value=cen["nome"])
        ws.cell(row=linha, column=2, value=cen["turno"])
        filtro = cen.get("filtro_candidatos") or []
        ws.cell(row=linha, column=3, value=", ".join(filtro) if filtro else "(todos)")
        linha += 1

    _ajustar_larguras(ws, {"A": 30, "B": 16, "C": 30})


def gerar_aba_resumo(wb, agregacoes_por_cenario):
    ws = wb.create_sheet("Resumo", 0)
    ws["A1"] = "Agregador de Pesquisas — Eleições 2026"
    ws["A1"].font = TITULO_FONT
    ws.merge_cells("A1:E1")

    ws["A3"] = "Resumo por cenário"
    ws["A3"].font = SUBTITULO_FONT

    linha = 5
    for nome_cenario, ag in agregacoes_por_cenario.items():
        ws.cell(row=linha, column=1, value=nome_cenario).font = Font(name="Arial", bold=True, size=12)
        ws.cell(row=linha, column=2, value=f"{ag['n_pesquisas']} pesquisas").font = Font(name="Arial", italic=True)
        linha += 1

        if ag["n_pesquisas"] == 0:
            ws.cell(row=linha, column=1, value="(sem dados)").font = Font(name="Arial", italic=True, color="888888")
            linha += 2
            continue

        for c in ag["candidatos"]:
            ws.cell(row=linha, column=1, value=c).font = Font(name="Arial")
            cell = ws.cell(row=linha, column=2, value=round(ag["medias"].get(c, 0), 2))
            cell.font = Font(name="Arial")
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="right")
            linha += 1
        linha += 1

    _ajustar_larguras(ws, {"A": 28, "B": 18})


def gerar_excel(
    caminho,
    agregacoes_por_cenario,
    todas_pesquisas,
    pesos_institutos,
    config,
    pesquisas_por_cenario,
    series_por_cenario=None,
):
    Path(caminho).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    gerar_aba_resumo(wb, agregacoes_por_cenario)

    for nome_cenario, ag in agregacoes_por_cenario.items():
        gerar_aba_cenario(
            wb, nome_cenario, ag,
            pesquisas_por_cenario.get(nome_cenario, []),
            pesos_institutos, config,
        )

    # Abas de evolução temporal (Fase A) — só para cenários que têm série
    if series_por_cenario:
        for nome_cenario, serie in series_por_cenario.items():
            if serie and serie.get("pontos"):
                gerar_aba_evolucao(wb, nome_cenario, serie)

    gerar_aba_todas_pesquisas(wb, todas_pesquisas)
    gerar_aba_config(wb, config, pesos_institutos)

    wb.save(caminho)
    return caminho
